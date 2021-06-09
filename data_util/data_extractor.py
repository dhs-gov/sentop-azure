import logging
import io
import nltk
nltk.download('punkt', download_dir='.')
#nltk.data.path.append("/")
from globals import globalutils
from openpyxl import load_workbook
from database import postgres
import urllib3
from openpyxl.utils import get_column_letter
from . import data_cleaner


# Return data, error, stop_words
def get_json_payload(json_obj):
    sentlog = globalutils.SentopLog()
    user_stop_words = []
    annotation = None
    try:
        user_stop_words = json_obj.get('user_stop_words')
        annotation = json_obj.get('annotation')

        docs = json_obj.get('documents')
        row_id_list = []
        data_list = []
        if docs:
            for doc in docs:
                row_id = doc.get('id')
                row_id_list.append(row_id)
                text = doc.get('text')
                if not text:
                    text = "NA"
                data_list.append(text)
        #else:
        #    sentlog.append(f"WARNING! No JSON documents found.<br>")
        return row_id_list, data_list, user_stop_words, annotation, None
    except Exception as e:    
        globalutils.show_stack_trace(str(e))
        return None, None, None, None, str(e)


def get_col_values(ws):
    sentlog = globalutils.SentopLog()
    try:
        params = []
        for row in ws.iter_rows():
            for col_cell in row:
                val = col_cell.value
                params.append(val)
        return params
    except Exception as e:
        globalutils.show_stack_trace(str(e))

def get_col_values_as_str(ws):
    sentlog = globalutils.SentopLog()

    try:
        annotation = ""
        for row in ws.iter_rows():
            for col_cell in row:
                val = col_cell.value
                annotation = annotation + val
        return annotation
    except Exception as e:
        globalutils.show_stack_trace(str(e))

def get_xlsx_data(bytes):
    sentlog = globalutils.SentopLog()
    try:
        xlsx = io.BytesIO(bytes)
        wb = load_workbook(xlsx)
        sentlog.append(f"<b>&#8226; Worksheets:</b><br>")
        sentlog.append("<pre>")
        for x in wb.sheetnames:
            sentlog.append(f"- {x}")
        sentlog.append("</pre>")
        ws = wb.worksheets[0]  

        # ------------------------ FIND DATA ID COLUMN -------------------------

        # NOTE: Selected data column header cell MUST NOT BE BLANK!
        
        #sentlog.append(f"----------------------------------------------------------")
        id_column = None
        row = ws[1]
        col_num = 0
        # Data ID column, if it exists, MUST have RED background color red (#FF0000) in FIRST HEADER ROW.
        #sentlog.append("Searching for data ID column (denoted by cell color #FF0000).")
        id_column = None
        for col_cell in row:
            col_num = col_num + 1
            col_letter = get_column_letter(col_num)
            cell_address = col_letter + str(1)
            #print(f"Cell address: {cell_address}")            
            color_in_hex = ws.cell(1, col_num).fill.fgColor.rgb
            #print(f"Color type: {type(color_in_hex)}")
            if (type(color_in_hex) == str):
                color_in_hex = str(color_in_hex)
                color_in_hex = color_in_hex[2:]
                #sentlog.append(f"color_in_hex: {color_in_hex}")
                if len(color_in_hex) == 6:
                    if color_in_hex == 'FF0000':
                        sentlog.append(f"<b>&#8226; Found cell color:</b> 'FF0000' at cell {cell_address}<br>")
                        id_column = col_letter
                        break

        if not id_column:
            sentlog.append("<div style=\"color: #e97e16; font-weight: bold; \">&#8226; Warning:  1 with color #FF0000 in XLSX file. Will use row number as ID.</div>")
        else:
            sentlog.append(f"<b>&#8226; Found data ID column:</b> {id_column}<br>")

        # --------------------- GET HIGHLIGHTED DATA CELLS ---------------------
        
        # NOTE: Cell color will be '000000' if cell has no data, regardless of color.
        
        #sentlog.append(f"----------------------------------------------------------")
        row_num = 0
        data_list = []
        data_id = None
        for row in ws.iter_rows():
            row_num = row_num + 1
            col_num = 0
            for col_cell in row:
                col_num = col_num + 1
                col_letter = get_column_letter(col_num)
                cell_address = col_letter + str(row_num)
                #print(f"Cell: {cell_address}")
                if not id_column or id_column == 'None':
                    data_id = row_num
                elif col_letter == id_column:
                    data_id = col_cell.value
                
                if col_letter != id_column:
                    color_in_hex = ws.cell(row_num, col_num).fill.fgColor.rgb

                    if type(color_in_hex) == str:
                        color_in_hex = color_in_hex[2:]
                        #print(f"{cell_address}: {color_in_hex}")
                        if color_in_hex == 'FFFF00':  # Yellow
                            print(f"Found data cell: {cell_address}")
                            val = col_cell.value

                            if not val:
                                val = "NA"
                            if val.find('_x000d_'):
                                val = val.replace('_x000d_', '')
                                val = val.replace('_x000D_', '')
                                val = val.replace('_X000D_', '')
                                #sentlog.append("Replaced '_x000d_'")
                                #sentlog.append(f"New text: {val}")
                            data_list.append([str(data_id) + '_' + col_letter, val])

     
        print("----------------------------")
        #for data in data_list:
        #    print(f"DATA: {data[0]}, {data[1]}")
        
        # Get column slice
        row_id_list = globalutils.column(data_list, 0)
        #for sentop_idz in row_id_list:
        #    print(f"sentop_id: {sentop_idz}")

        # Get column slice
        main_data_list = globalutils.column(data_list, 1)
        #for data in main_data_list:
        #    print(f"MAIN: {data}")

        # --------------------- GET STOP WORDS ---------------------

        # Stop words must be included as a sheet in the XLSX file.
        user_stop_words = []
        annotation = None
        try:
            stop_words_ws = wb.get_sheet_by_name('SENTOP Stop Words')
            if stop_words_ws:
                user_stop_words = get_col_values(stop_words_ws)
                sentlog.append(f"<b>&#8226; User stop words:</b><br>")
                sentlog.append("<pre>")
                for x in user_stop_words:
                    sentlog.append(f"- {x}")
                sentlog.append("</pre>")
            else:
                #sentlog.append(f"No user stop words found.")
                user_stop_words = []

        # --------------------- GET ANNOTATION ---------------------

            annotation_ws = wb.get_sheet_by_name('SENTOP Annotation')
            if annotation_ws:
                annotation = get_col_values_as_str(annotation_ws)
                #sentlog.append(f"Found user annotation: {annotation}")
            else:
                sentlog.append(f"No user annotation found.")
                annotation = None
        except Exception as e:
            globalutils.show_stack_trace(str(e))
            user_stop_words = [] 
            annotation = None

        if row_id_list and main_data_list:
            #print(f"Returning data_list len: {len(main_data_list)}")
            return row_id_list, main_data_list, user_stop_words, annotation, None
        else:
            print("row_id_list or data_list is None!")
            return None, None, None, None, "Could not extract XLSX data."

    except Exception as e:   
        globalutils.show_stack_trace(str(e))
        return None, None, None, None, str(e)


# ================================= M A I N ===================================

# Return data[], user_stop_words[], error
def get_data(req, sentop_id, kms_id):
    sentlog = globalutils.SentopLog()

    try:
        # -------------------------- GET JSON DATA -----------------------------
        # Check JSON payload even if kms_id since user_stop_words may exists.
        #sentlog.append("Checking for JSON data.")
        body_bytes = req.get_body()
        row_id_list = []
        user_stop_words = []
        annotation = None
        data_list = []
        error = None

        if body_bytes:
            body_str = str(body_bytes, 'utf-8')
            body_str_cleaned = data_cleaner.clean_postgres_json(body_str)
            db = postgres.Database()
            error = db.add_json_data(sentop_id, body_str_cleaned)
            if error:
                return None, None, None, error

            json_obj = req.get_json()
            if json_obj:
                # NOTE: data is list of [sentop_id,text]
                row_id_list, data_list, user_stop_words, annotation, error = get_json_payload(json_obj)
                if error:
                    return None, None, None, None, error
            else:
                print("No JSON object found in request.")
        else:
            print("No JSON body found in request")

        if data_list:
            sentlog.append("<b>&#8226; Data type:</b> JSON<br>")
            return row_id_list, data_list, user_stop_words, annotation, error
        elif not kms_id:
            return None, None,  None, None, "No JSON or file URL received."

        # ------------------------ GET FILE DATA ---------------------------

        sentlog.append("<div style=\"color: #e97e16; font-weight: bold; \">&#8226; Warning: No JSON payload found.</div>")
        sentlog.append(f"<b>&#8226; File URL:</b> {kms_id}.<br>") 
        http = urllib3.PoolManager()
        # NOTE: Use http://, not file://, scheme -- The file must be
        # located where the HTTP server is started. Note that the
        # response data will always be binary.
        resp = http.request('GET', kms_id)

        error = None
        if resp.status == 200:
            sentlog.append(f"<b>&#8226; File found:</b> True<br>")
            try:

                # --------------------- GET XLSX DATA ----------------------

                if kms_id.endswith('.xlsx'):
                    #sentlog.append("<b>&#8226; File type:</b> XLSX<br>")
                    try:
                        row_id_list, data_list, user_stop_words, annotation, error = get_xlsx_data(resp.data)
                        if data_list:
                            sentlog.append("<b>&#8226; Data type:</b> XLSX<br>")

                        if error:
                            return None, None, None, None, error
                    except Exception as e:
                        globalutils.show_stack_trace(str(e))
                        return None, None, None, None, str(e)

                else:
                    sentlog.append("Error: File extension not supported: ", kms_id)
                    return None, None, None, None, "File extension not supported."

                if data_list:
                    #sentlog.append("Extracted XLSX data.<br>")
                    return row_id_list, data_list, user_stop_words, annotation, None
                else:
                    return None, None, None, None, "Could not extract XLSX data."

            except Exception as e:    
                globalutils.show_stack_trace(str(e))
                return None, None, None, None, str(e)
        else:
            return None, None,  None, None, "ERROR: Could not find file."

    except Exception as e:    
        globalutils.show_stack_trace(str(e))
        return None, None,  None, None, str(e)
        
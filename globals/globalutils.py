import os
from contextlib import contextmanager,redirect_stderr,redirect_stdout
from os import devnull
import sys
import logging
import psutil
from sklearn.feature_extraction import text
from topic_modeling import stopwords
from . import globalvars
import sentop_config as config
import nltk
from nltk.tokenize import word_tokenize
import string
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
from dateutil import tz

class Sentiments:
    def __init__(self, id, name, model_name, type, data_list):
        self.id = id
        self.name = name
        self.model_name = model_name
        self.type = type
        self.data_list = data_list

def column(matrix, i):
    return [row[i] for row in matrix]


class TextChecker():
    def __init__(self, all_stop_words):
        self.all_stop_words = all_stop_words

    # NOTE: text is case sensitive.
    def check_entire_text(self, text):
        if text in self.all_stop_words:
            return None
        else:
            return text

    # NOTE: text is case insensitive
    def check_each_word(self, text):
        # Remove punctuation
        text = text.translate(str.maketrans('', '', string.punctuation))
        words = text.split()
        new_text = ""
        for word in words:
            if word.lower() not in self.all_stop_words :
                #print(f"New word: {word}")
                return word
            #else:
            #    print(f"Bad word: {word}")
        return None

    # Counts valid number of words (i.e., words that contain at least one alpha char)
    def check_num_words(self, text):
        words = text.split()
        num_good_words = 0
        for word in words:
            found_letter = any(c.isalpha() for c in word)
            if found_letter:
                num_good_words = num_good_words + 1
            #else:
            #    print(f"Bad: {word}")
        if num_good_words >= globalvars.MIN_DOC_WORDS:
            return True
        else:
            return False

    def trim_text(self, text):
        # Make sure text is not over maximum number of embedding tokens
        if len(nltk.word_tokenize(text)) > globalvars.MAX_TOKENS:
            # Get only the last sentence
            text = text[0:globalvars.MAX_TOKENS]
        return text




@contextmanager
def suppress_stdout_stderr():
    """A context manager that redirects stdout and stderr to devnull"""
    with open(devnull, 'w') as fnull:
        with redirect_stderr(fnull) as err, redirect_stdout(fnull) as out:
            yield (err, out)

# Disable logging
def block_logging():  
    # Disable logging  
    logging.config.dictConfig({
        'version': 1,
        'disable_existing_loggers': True,
    })


# Restore logging
def enable_logging():
    # Re-enable logging
    logging.config.dictConfig({
        'version': 1,
        'disable_existing_loggers': False,
    })


def get_memory():
    memory = psutil.Process(os.getpid()).memory_info().rss / 1024 ** 2
    mem_str = str(memory) + "MB"
    return mem_str


def get_memory():
    memory = psutil.Process(os.getpid()).memory_info().rss / 1024 ** 2
    mem_str = str(memory) + "MB"
    return mem_str


def get_stopwords_list(user_stop_words):
    frozen_stopwords = get_frozen_stopwords(user_stop_words)
    return list(frozen_stopwords)


# NOTE: user_stop_words are case sensitive. All other stopwords are case 
# insensitive.
def get_frozen_stopwords(user_stop_words):
    lowercase_sentop_stopwords = [x.lower() for x in stopwords.stopwords_list]
    lowercase_sentop_stopwords.extend(user_stop_words)
    all_stop_words = text.ENGLISH_STOP_WORDS.union(lowercase_sentop_stopwords)
    # print("all_stop_words: ", all_stop_words)
    return all_stop_words


def get_sentiment(id, sentiments):
    for r in sentiments:
        if r.id == id:
            print(f"Found ID: {r.id}")
            return r


def generate_excel(id, annotation, num_list, data_list, sentiment_results, bertopic_results, lda_results):
   
    sentlog = sentop_log.SentopLog()
    try:
        bert_sentence_topics = bertopic_results.topic_per_row
        bert_topics = bertopic_results.topics_list
        bert_duplicate_words = bertopic_results.duplicate_words_across_topics

        lda_sentence_topics = lda_results.topic_per_row
        lda_topics = lda_results.topics_list
        lda_duplicate_words = lda_results.duplicate_words_across_topics

        output_dir_path = config.data_dir_path.get("output")

        class3 = get_sentiment('class3', sentiment_results)
        class5 = get_sentiment('class5', sentiment_results)
        emotion1 = get_sentiment('emotion1', sentiment_results)
        offensive1 = get_sentiment('offensive1', sentiment_results) 
        emotion2 = get_sentiment('emotion2', sentiment_results)

        # Create results data
        rows = []
        for i in range(len(data_list)):
            row_data = []
            if (data_list[i]):
                row_data.append(num_list[i])
                row_data.append(data_list[i])
                row_data.append(bert_sentence_topics[i])
                row_data.append(lda_sentence_topics[i])
                row_data.append(class3.data_list[i])
                row_data.append(class5.data_list[i])
                row_data.append(emotion1.data_list[i])
                row_data.append(emotion2.data_list[i])
                row_data.append(offensive1.data_list[i])
                rows.append(row_data)

        # Create results XLSX
        wb = Workbook()
        xlsx_out = output_dir_path + "\\" + id + "_results.xlsx"
        ws1 = wb.active
        ws1.title = "Results"
        ws1.append(['ID', 'Document', 'BERTopic Topic', 'LDA Topic', class3.name, class5.name, emotion1.name, emotion2.name, offensive1.name])
        ws1['A1'].font = Font(bold=True)
        ws1['B1'].font = Font(bold=True)
        # Topic columns
        ws1['C1'].fill = PatternFill(start_color='FF66FF66', end_color='FF66FF66', fill_type='solid')
        ws1['C1'].font = Font(bold=True)
        ws1['D1'].fill = PatternFill(start_color='FF66FF66', end_color='FF66FF66', fill_type='solid')
        ws1['D1'].font = Font(bold=True)

        # Polarity sentiment columns
        ws1['E1'].fill = PatternFill(start_color='FF66FFFF', end_color='FF66FFFF', fill_type='solid')
        ws1['E1'].font = Font(bold=True)
        ws1['F1'].fill = PatternFill(start_color='FF66FFFF', end_color='FF66FFFF', fill_type='solid')
        ws1['F1'].font = Font(bold=True)

        # Emotion sentiment columns
        ws1['G1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['G1'].font = Font(bold=True)
        ws1['H1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['H1'].font = Font(bold=True)
        ws1['I1'].fill = PatternFill(start_color='FFFFFF99', end_color='FFFFFF99', fill_type='solid')
        ws1['I1'].font = Font(bold=True)

        for i in range(len(rows)):
            ws1.append(rows[i])

        # Create Annotation XLSX sheet
        ws4 = wb.create_sheet(title="Annotation")
        fields = ['Annotation']
        annotation_list = []
        annotation_list.append(annotation)
        ws4.append(annotation_list)

        # Create BERTopic topics data
        rows = []
        for i in range(len(bert_topics)):
            for j in range(len(bert_topics[i].words)):
                row_data = []
                row_data.append(bert_topics[i].topic_num)
                row_data.append(bert_topics[i].words[j])
                row_data.append(float(bert_topics[i].weights[j]))
                rows.append(row_data)

        # Create BERTopic topics data XLSX sheet
        ws2 = wb.create_sheet(title="BERTopic")
        ws2.append(['Topic', 'Top Words', 'Weight'])
        for i in range(len(rows)):
            ws2.append(rows[i])

        # Create BERTopic non-overlapping topic words data
        rows = []
        for i in range(len(bert_topics)):
            for j in range(len(bert_topics[i].words)):
                if not bert_topics[i].words[j] in bert_duplicate_words:
                    row_data = []
                    row_data.append(bert_topics[i].topic_num)
                    row_data.append(bert_topics[i].words[j])
                    row_data.append(float(bert_topics[i].weights[j]))
                    rows.append(row_data)

        # Create BERTopic non-overlapping topics data XLSX sheet
        ws2 = wb.create_sheet(title="BERTopic Non-Overlapping Topics")
        ws2.append(['Topic', 'Top Words', 'Weight'])
        for i in range(len(rows)):
            ws2.append(rows[i])  

        # Create LDA topics data
        rows = []
        for i in range(len(lda_topics)):
            #print("LDA i: ", i)
            for j in range(len(lda_topics[i].words)):
                #print("LDA j: ", j)
                row_data = []
                row_data.append(lda_topics[i].topic_num)
                row_data.append(lda_topics[i].words[j])
                row_data.append(float(lda_topics[i].weights[j]))
                rows.append(row_data)

        # Create LDA topics data XLSX sheet
        ws3 = wb.create_sheet(title="LDA")
        fields = ['Topic', 'Top Words', 'Weight']
        ws3.append(fields)
        for i in range(len(rows)):
            ws3.append(rows[i])

        # Create LDA non-overlapping topics words data
        rows = []
        for i in range(len(lda_topics)):
            #print("LDA i: ", i)
            for j in range(len(lda_topics[i].words)):
                if not lda_topics[i].words[j] in lda_duplicate_words:
                    row_data = []
                    row_data.append(lda_topics[i].topic_num)
                    row_data.append(lda_topics[i].words[j])
                    row_data.append(float(lda_topics[i].weights[j]))
                    rows.append(row_data)

        # Create LDA topics data XLSX sheet
        ws3 = wb.create_sheet(title="LDA Non-Overlapping Topics")
        fields = ['Topic', 'Top Words', 'Weight']
        ws3.append(fields)
        for i in range(len(rows)):
            ws3.append(rows[i])

        # Save XLSX
        wb.save(filename = xlsx_out)
    except Exception as e:
        show_stack_trace(e)

